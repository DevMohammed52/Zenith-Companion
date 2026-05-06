#!/usr/bin/env python3
"""Extract local IdleMMO pet workbook data into gitignored JSON files.

This script does not call the IdleMMO API. It reads the two workbook files in
game_info/05_pets and writes normalized research extracts under local_data/.
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Use the Codex bundled Python runtime or install openpyxl locally."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "game_info" / "05_pets"
DEFAULT_OUTPUT_DIR = ROOT / "local_data" / "derived_manual" / "pets"

CHAPTER_WORKBOOK = "Chapter II - Pets.xlsx"
MASTER_WORKBOOK = "IdleMMO Pet Battles - Master Sheet (Public).xlsx"
MEGA_TEST_WORKBOOK = "IdleMMO Pet Battles - Mega Test.xlsx"
PET_VALUE_WORKBOOK = "IMMO - Pet Value Calculator.xlsx"

ZONE_ALIAS = {
    "Lvl 1 - BB": "Level 1 - Bluebell Hollow",
    "Level 1 Zone": "Level 1 - Bluebell Hollow",
    "Lvl 8 - WW": "Level  8 - Whispering Woods",
    "Level 8 Zone": "Level  8 - Whispering Woods",
    "Lvl 18 - EL": "Level 18 - Eldoria",
    "Level 18 Zone": "Level 18 - Eldoria",
    "Lvl 32 - CC": "Level 32 - Crystal Caverns",
    "Level 32 Zone": "Level 32 - Crystal Caverns",
    "Lvl 48 - SP": "Level 48 - Skyreach Peak",
    "Level 48 Zone": "Level 48 - Skyreach Peak",
    "Lvl 60 - EO": "Level 60 - Enchanted Oasis",
    "Level 60 Zone": "Level 60 - Enchanted Oasis",
    "Lvl 70 - FG": "Level 70 - Floating Gardens of Aetheria",
    "Level 70 Zone": "Level 70 - Floating Gardens of Aetheria",
    "Lvl 78 - CO": "Level 78 - Celestial Observatory",
    "Level 78 Zone": "Level 78 - Celestial Observatory",
    "Lvl 92 - IW": "Level 92 - Isle of Whispers",
    "Level 92 Zone": "Level 92 - Isle of Whispers",
    "Lvl 100 - CI": "Level 100 - Citadel",
    "Level 100 Zone": "Level 100 - Citadel",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def number(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return int(parsed) if parsed.is_integer() else parsed


def seconds(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, timedelta):
        total = value.total_seconds()
        return int(total) if total.is_integer() else total
    if isinstance(value, (int, float)):
        return int(value * 86400)
    return None


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, Path):
        return str(value)
    return str(value)


def cell_payload(formula_value: Any, cached_value: Any) -> dict[str, Any] | None:
    """Return a compact serializable cell payload for a non-empty workbook cell."""
    if formula_value is None and cached_value is None:
        return None
    payload: dict[str, Any] = {}
    if isinstance(formula_value, str) and formula_value.startswith("="):
        payload["formula"] = formula_value
        if cached_value is not None:
            payload["value"] = json_safe(cached_value)
    else:
        payload["value"] = json_safe(formula_value)
    return payload


def dump_workbook_raw(path: Path, output_dir: Path) -> dict[str, Any]:
    """Dump every non-empty cell from a workbook for audit/re-extraction."""
    wb_formula = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    wb_cached = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    workbook_slug = path.stem.lower().replace(" ", "_").replace("-", "_")
    sheets = []
    for ws in wb_formula.worksheets:
        cached_ws = wb_cached[ws.title]
        cells = []
        cached_rows = cached_ws.iter_rows()
        for row_index, formula_row in enumerate(ws.iter_rows(), start=1):
            try:
                cached_row = next(cached_rows)
            except StopIteration:
                cached_row = []
            for col_index, formula_cell in enumerate(formula_row, start=1):
                formula_value = formula_cell.value
                cached_value = cached_row[col_index - 1].value if col_index <= len(cached_row) else None
                payload = cell_payload(formula_value, cached_value)
                if payload:
                    payload["address"] = formula_cell.coordinate
                    payload["row"] = row_index
                    payload["col"] = col_index
                    cells.append(payload)
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": [],
                "non_empty_cells": len(cells),
                "cells": cells,
            }
        )
    payload = {
        "source_file": str(path.relative_to(ROOT)),
        "generated_at": now_iso(),
        "sheets": sheets,
    }
    write_json(output_dir / "workbooks" / f"{workbook_slug}.json", payload)
    return {
        "source_file": str(path.relative_to(ROOT)),
        "raw_dump_file": str((output_dir / "workbooks" / f"{workbook_slug}.json").relative_to(ROOT)),
        "sheets": [
            {
                "name": sheet["name"],
                "max_row": sheet["max_row"],
                "max_column": sheet["max_column"],
                "non_empty_cells": sheet["non_empty_cells"],
            }
            for sheet in sheets
        ],
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_safe),
        encoding="utf-8",
    )
    temp_path.replace(path)


def stat_pair(ws: Any, row: int, base_col: int, per_level_col: int) -> dict[str, int | float | None]:
    return {
        "base": number(ws.cell(row, base_col).value),
        "per_level": number(ws.cell(row, per_level_col).value),
    }


def split_egg_text(value: Any) -> dict[str, Any] | None:
    text = clean_string(value)
    if not text:
        return None
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    name = lines[0] if lines else text
    chance = None
    if len(lines) > 1:
        chance_text = lines[1].replace("~", "").replace("%", "").strip()
        try:
            chance = float(chance_text)
        except ValueError:
            chance = None
    return {"egg": name, "chance_percent": chance, "raw": text}


def extract_chapter_workbook(input_dir: Path) -> dict[str, Any]:
    path = input_dir / CHAPTER_WORKBOOK
    wb_values = load_workbook(path, read_only=True, data_only=True, keep_links=False)

    pet_stats_ws = wb_values["🦊 Pet stats"]
    species = []
    for row in range(3, 53):
        name = clean_string(pet_stats_ws.cell(row, 6).value)
        quality = clean_string(pet_stats_ws.cell(row, 2).value)
        if not name:
            continue
        species.append(
            {
                "name": name,
                "quality": quality,
                "stats": {
                    "agility": stat_pair(pet_stats_ws, row, 7, 8),
                    "accuracy": stat_pair(pet_stats_ws, row, 10, 11),
                    "protection": stat_pair(pet_stats_ws, row, 13, 14),
                    "attack_power": stat_pair(pet_stats_ws, row, 16, 17),
                    "movement_speed": stat_pair(pet_stats_ws, row, 19, 20),
                    "max_health": stat_pair(pet_stats_ws, row, 22, 23),
                    "max_stamina": stat_pair(pet_stats_ws, row, 25, 26),
                    "critical_damage": stat_pair(pet_stats_ws, row, 28, 29),
                    "critical_chance": stat_pair(pet_stats_ws, row, 31, 32),
                },
            }
        )

    acquiring_ws = wb_values["🐉 Acquiring pets"]
    acquisition_qualities = [
        ("standard", 5, 6),
        ("refined", 7, 8),
        ("premium", 9, 10),
        ("epic", 11, 12),
        ("legendary", 13, 14),
        ("mythic", 15, 16),
    ]
    acquisition_rows = []
    for row in range(3, 48):
        location = clean_string(acquiring_ws.cell(row, 1).value)
        boss = clean_string(acquiring_ws.cell(row, 2).value)
        if not location and not boss:
            continue
        eggs = []
        for quality, _image_col, text_col in acquisition_qualities:
            egg = split_egg_text(acquiring_ws.cell(row, text_col).value)
            if egg:
                egg["quality"] = quality
                eggs.append(egg)
        acquisition_rows.append(
            {
                "location": location,
                "boss": boss,
                "level_requirement": clean_string(acquiring_ws.cell(row, 4).value),
                "eggs": eggs,
            }
        )

    rarity_ws = wb_values["👑 Pet Rarity"]
    rarity_rows = []
    for row in range(4, 35):
        pet_name = clean_string(rarity_ws.cell(row, 4).value)
        if not pet_name:
            continue
        rarity_rows.append(
            {
                "id": clean_string(rarity_ws.cell(row, 1).value),
                "quality": clean_string(rarity_ws.cell(row, 2).value),
                "pet_name": pet_name,
                "world_boss": clean_string(rarity_ws.cell(row, 5).value),
                "battle_plus_respawn_seconds": seconds(rarity_ws.cell(row, 6).value),
                "drop_chance_percent": number(rarity_ws.cell(row, 7).value),
            }
        )

    data_ws = wb_values["Data"]
    mastery_rows = []
    for row in range(3, 103):
        level = number(data_ws.cell(row, 2).value)
        if level is None:
            continue
        mastery_rows.append(
            {
                "level": level,
                "stat_bonus_percent": number(data_ws.cell(row, 3).value),
                "concurrent_battles": number(data_ws.cell(row, 4).value),
                "loot_chance": number(data_ws.cell(row, 5).value),
            }
        )

    concurrent_slots = []
    for row in range(5, 16):
        required_level = number(data_ws.cell(row, 7).value)
        slots = number(data_ws.cell(row, 8).value)
        if required_level is not None and slots is not None:
            concurrent_slots.append(
                {"required_pet_mastery_level": required_level, "slots": slots}
            )

    loot_chance_breakpoints = []
    for row in range(5, 16):
        level = number(data_ws.cell(row, 10).value)
        loot_chance = number(data_ws.cell(row, 11).value)
        if level is not None and loot_chance is not None:
            loot_chance_breakpoints.append({"level": level, "loot_chance": loot_chance})

    stamina_drain_by_quality = []
    for row in range(5, 11):
        quality = clean_string(data_ws.cell(row, 14).value)
        drain = number(data_ws.cell(row, 15).value)
        if quality:
            stamina_drain_by_quality.append(
                {"quality": quality, "battle_stamina_per_second": drain}
            )

    return {
        "source_file": str(path.relative_to(ROOT)),
        "species": species,
        "acquisition": acquisition_rows,
        "rarity": rarity_rows,
        "pet_mastery": mastery_rows,
        "concurrent_slots": concurrent_slots,
        "loot_chance_breakpoints": loot_chance_breakpoints,
        "stamina_drain_by_quality": stamina_drain_by_quality,
        "known_formulas": {
            "raw_stat": "base + ((pet_level - 1) * per_level)",
            "hunting_time_per_enemy_days": "(200 - 125 * (0.7 * MIN(agility / 120, 1) + 0.3 * MIN(movement_speed / 100, 1))) / 86400",
            "pet_mastery_loot_chance": "0.025 + pet_mastery_level * 0.00075",
            "concurrent_slot_required_level": "round(1 + (slot_count - 2) * 12.5, 0) - 1",
        },
    }


def extract_master_workbook(input_dir: Path) -> dict[str, Any]:
    path = input_dir / MASTER_WORKBOOK
    wb_values = load_workbook(path, read_only=True, data_only=True, keep_links=False)

    all_pet_ws = wb_values["All Pet Data"]
    zone_starts = [5, 37, 69, 101, 133, 165, 197, 229, 261, 293]
    zone_records = []
    for start in zone_starts:
        zone_label = clean_string(all_pet_ws.cell(start, 2).value)
        if not zone_label:
            continue
        pets = []
        for row in range(start + 2, start + 32):
            pet = clean_string(all_pet_ws.cell(row, 2).value)
            if not pet:
                continue
            pets.append(
                {
                    "pet_name": pet,
                    "battle_time_seconds": seconds(all_pet_ws.cell(row, 3).value),
                    "enemies_battled": number(all_pet_ws.cell(row, 4).value),
                    "loot_pieces": number(all_pet_ws.cell(row, 5).value),
                    "expected_revenue_per_battle": number(all_pet_ws.cell(row, 7).value),
                    "expected_revenue_per_hour": number(all_pet_ws.cell(row, 8).value),
                    "expected_profit_per_battle": number(all_pet_ws.cell(row, 9).value),
                    "expected_profit_per_hour_no_sleep": number(all_pet_ws.cell(row, 10).value),
                    "expected_profit_per_hour_with_sleep": number(all_pet_ws.cell(row, 11).value),
                    "expected_profit_per_hour_healing_with_sleep": number(all_pet_ws.cell(row, 12).value),
                    "profit_margin": number(all_pet_ws.cell(row, 13).value),
                    "food_cost_per_hour_cheapest": number(all_pet_ws.cell(row, 14).value),
                    "max_stamina": number(all_pet_ws.cell(row, 20).value),
                    "stamina_drain_per_hour": number(all_pet_ws.cell(row, 21).value),
                    "stamina_drain_per_battle": number(all_pet_ws.cell(row, 22).value),
                    "time_battled_for_zero_stamina_seconds": seconds(all_pet_ws.cell(row, 23).value),
                    "stamina_recovery_zero_to_full_seconds": seconds(all_pet_ws.cell(row, 24).value),
                    "battles_before_sleep": number(all_pet_ws.cell(row, 25).value),
                    "health_recovery_zero_to_full_seconds": seconds(all_pet_ws.cell(row, 26).value),
                    "sleep_to_battle_for_stamina": number(all_pet_ws.cell(row, 27).value),
                    "battle_to_sleep_for_hp": number(all_pet_ws.cell(row, 28).value),
                }
            )
        zone_records.append({"zone": zone_label, "pets": pets})

    prices_ws = wb_values["Prices"]
    prices = []
    current_zone = None
    for row in range(4, 123):
        zone = clean_string(prices_ws.cell(row, 2).value)
        if zone:
            current_zone = zone
        loot_name = clean_string(prices_ws.cell(row, 3).value)
        if loot_name:
            prices.append(
                {
                    "zone": current_zone,
                    "loot_name": loot_name,
                    "market_price": number(prices_ws.cell(row, 4).value),
                    "last_day_price_after_tax": number(prices_ws.cell(row, 5).value),
                    "vendor_100_barter": number(prices_ws.cell(row, 6).value),
                    "best_after_tax_sell_value": number(prices_ws.cell(row, 7).value),
                }
            )

    food_prices = []
    for row in range(4, 20):
        food = clean_string(prices_ws.cell(row, 9).value)
        price = number(prices_ws.cell(row, 10).value)
        if food:
            food_prices.append({"food": food, "price": price})

    drops_ws = wb_values["Expected Drop %"]
    expected_drops = []
    current_zone = None
    for row in range(5, 159):
        zone = clean_string(drops_ws.cell(row, 11).value)
        if zone and not zone.startswith("*"):
            current_zone = zone
        item = clean_string(drops_ws.cell(row, 2).value)
        if item:
            expected_drops.append(
                {
                    "zone": current_zone,
                    "item_name": item,
                    "expected_drop_percent": number(drops_ws.cell(row, 5).value),
                    "drop_value_share_max_price": number(drops_ws.cell(row, 6).value),
                    "drop_value_share_vendor": number(drops_ws.cell(row, 7).value),
                    "value_share_max_price": number(drops_ws.cell(row, 8).value),
                }
            )

    return {
        "source_file": str(path.relative_to(ROOT)),
        "zone_pet_battle_records": zone_records,
        "prices": prices,
        "food_prices": food_prices,
        "expected_drops": expected_drops,
    }


def extract_mega_test_workbook(input_dir: Path) -> dict[str, Any]:
    path = input_dir / MEGA_TEST_WORKBOOK
    if not path.exists():
        return {"source_file": str(path.relative_to(ROOT)), "missing": True}
    wb_values = load_workbook(path, read_only=False, data_only=True, keep_links=False)

    result_ws = wb_values["RESULT OF TESTS"]
    zone_rankings = []
    for row in range(4, 13):
        zone = clean_string(result_ws.cell(row, 3).value)
        if not zone:
            continue
        zone_rankings.append(
            {
                "rank": clean_string(result_ws.cell(row, 2).value),
                "zone": zone,
                "normalized_zone": ZONE_ALIAS.get(zone, zone),
                "profit_per_hour_pm100": number(result_ws.cell(row, 4).value),
                "pet_exp_profit_efficiency_scale": number(result_ws.cell(row, 5).value),
            }
        )

    result_sections = []
    headers = [
        "pet",
        "total_revenue",
        "total_cost",
        "total_profit",
        "total_time_seconds",
        "revenue_per_hour",
        "profit_per_hour",
        "profit_margin",
        "total_loot",
        "total_pet_exp",
        "total_pet_mastery_exp",
        "pet_exp_per_hour",
        "pet_exp_per_profit",
        "time_per_loot_seconds",
    ]
    for row in range(17, (result_ws.max_row or 0) + 1):
        for col in (2, 17, 27, 40, 53, 66):
            label = clean_string(result_ws.cell(row, col).value)
            next_value = clean_string(result_ws.cell(row + 1, col).value) if row + 1 <= result_ws.max_row else None
            if label and next_value == "Pet":
                section_rows = []
                cursor = row + 2
                while cursor <= result_ws.max_row:
                    pet = clean_string(result_ws.cell(cursor, col).value)
                    if not pet or pet in {"Revenue = Price AVG", "PPH = Price P hour"}:
                        break
                    record = {"source_row": cursor}
                    for offset, key in enumerate(headers):
                        value = result_ws.cell(cursor, col + offset).value
                        record[key] = seconds(value) if key.endswith("_seconds") else json_safe(value)
                    section_rows.append(record)
                    cursor += 1
                result_sections.append(
                    {
                        "label": label,
                        "start_cell": result_ws.cell(row, col).coordinate,
                        "rows": section_rows,
                    }
                )

    calculator_ws = wb_values["CALCULATOR"]
    calculator_drop_breakdowns = []
    current_zone = None
    for row in range(1, (calculator_ws.max_row or 0) + 1):
        zone_label = clean_string(calculator_ws.cell(row, 7).value)
        if zone_label and zone_label.lower().startswith("level") and "zone" in zone_label.lower():
            current_zone = zone_label
            continue
        item = clean_string(calculator_ws.cell(row, 7).value)
        if not item or item in {"Material", "TOTAL"} or current_zone is None:
            continue
        calculator_drop_breakdowns.append(
            {
                "zone": current_zone,
                "normalized_zone": ZONE_ALIAS.get(current_zone, current_zone),
                "item_name": item,
                "price": number(calculator_ws.cell(row, 8).value),
                "total_drops": number(calculator_ws.cell(row, 9).value),
                "drop_percent": number(calculator_ws.cell(row, 10).value),
                "drop_value_share": number(calculator_ws.cell(row, 11).value),
                "value_share_percent": number(calculator_ws.cell(row, 12).value),
            }
        )

    prices_ws = wb_values["Loot Prices & Food costs"]
    loot_prices = []
    for row in range(4, 97):
        item = clean_string(prices_ws.cell(row, 2).value)
        if item:
            loot_prices.append(
                {
                    "item_name": item,
                    "last_day_price": number(prices_ws.cell(row, 3).value),
                    "last_day_price_after_market_tax": number(prices_ws.cell(row, 4).value),
                    "manual_price": number(prices_ws.cell(row, 5).value),
                    "manual_price_after_market_tax": number(prices_ws.cell(row, 6).value),
                }
            )

    food_prices = []
    for col in range(9, 20):
        food = clean_string(prices_ws.cell(48, col).value)
        price = number(prices_ws.cell(49, col).value)
        if food:
            food_prices.append({"food": food, "price": price})

    test_sheets = []
    for ws in wb_values.worksheets:
        if ws.title in {"RESULT OF TESTS", "CALCULATOR", "Loot Prices & Food costs"}:
            continue
        non_empty = 0
        for row in ws.iter_rows(values_only=True):
            if any(value is not None for value in row):
                non_empty += 1
        test_sheets.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty_rows": non_empty,
            }
        )

    return {
        "source_file": str(path.relative_to(ROOT)),
        "zone_rankings": zone_rankings,
        "result_sections": result_sections,
        "calculator_drop_breakdowns": calculator_drop_breakdowns,
        "loot_prices": loot_prices,
        "food_prices": food_prices,
        "test_sheets": test_sheets,
    }


def extract_pet_value_workbook(input_dir: Path) -> dict[str, Any]:
    path = input_dir / PET_VALUE_WORKBOOK
    if not path.exists():
        return {"source_file": str(path.relative_to(ROOT)), "missing": True}
    wb_values = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    wb_formula = load_workbook(path, read_only=False, data_only=False, keep_links=False)

    accounting_ws = wb_values["PET Accounting"]
    pet_accounting = []
    for row in range(5, 187):
        pet = clean_string(accounting_ws.cell(row, 3).value)
        if not pet:
            continue
        pet_accounting.append(
            {
                "pet_name": pet,
                "level": number(accounting_ws.cell(row, 4).value),
                "rarity": clean_string(accounting_ws.cell(row, 5).value),
                "egg_price": number(accounting_ws.cell(row, 6).value),
                "rough_estimate": number(accounting_ws.cell(row, 7).value),
            }
        )

    egg_price_table = []
    for row in range(4, 22):
        pet = clean_string(accounting_ws.cell(row, 15).value)
        price = number(accounting_ws.cell(row, 16).value)
        if pet:
            egg_price_table.append({"pet_name": pet, "egg_price": price})

    formula_text = clean_string(wb_formula["PET Accounting"].cell(5, 7).value) or ""
    level_100_bonus = []
    if "SWITCH" in formula_text:
        import re

        for name, bonus in re.findall(r'"([^"]+)"\s*,\s*([0-9]+)', formula_text):
            level_100_bonus.append({"pet_name": name, "level_100_bonus": int(bonus)})

    level_ws = wb_values["LEVELING 1 - 100"]
    leveling = []
    for row in range(2, 102):
        label = clean_string(level_ws.cell(row, 1).value)
        if not label:
            continue
        leveling.append(
            {
                "level_label": label,
                "level": number(str(label).replace("lvl", "").strip()),
                "exp_needed": number(level_ws.cell(row, 2).value),
                "total_exp": number(level_ws.cell(row, 3).value),
                "cumulative_percent": number(level_ws.cell(row, 4).value),
            }
        )

    return {
        "source_file": str(path.relative_to(ROOT)),
        "pet_accounting": pet_accounting,
        "egg_price_table": egg_price_table,
        "level_100_bonus_table": level_100_bonus,
        "leveling": leveling,
        "valuation_formula": {
            "description": "Rough estimate uses egg price plus weighted level-100 bonus. Weight = 85% cumulative EXP share + 15% linear level share.",
            "max_exp": 16310091,
        },
    }


def build_manifest(
    chapter: dict[str, Any],
    master: dict[str, Any],
    mega: dict[str, Any],
    value_calc: dict[str, Any],
    raw_summaries: list[dict[str, Any]],
    output_dir: Path,
) -> dict[str, Any]:
    battle_records = sum(len(zone["pets"]) for zone in master["zone_pet_battle_records"])
    return {
        "generated_at": now_iso(),
        "output_dir": str(output_dir.relative_to(ROOT)),
        "files": {
            "pet_species_stats.json": len(chapter["species"]),
            "pet_acquisition.json": len(chapter["acquisition"]),
            "pet_rarity.json": len(chapter["rarity"]),
            "pet_mastery.json": len(chapter["pet_mastery"]),
            "pet_battle_zones.json": battle_records,
            "pet_battle_prices.json": len(master["prices"]),
            "pet_battle_expected_drops.json": len(master["expected_drops"]),
            "pet_battle_mega_test.json": len(mega.get("calculator_drop_breakdowns", [])),
            "pet_value_calculator.json": len(value_calc.get("pet_accounting", [])),
        },
        "raw_workbook_dumps": raw_summaries,
        "notes": [
            "Generated from local workbook files only; no API calls.",
            "Keep generated output local until formulas are verified against in-game screenshots.",
            "Excel cannot evaluate the source workbook's Google Sheets finalValue function locally.",
            "Raw workbook dumps include every non-empty cell from each workbook for audit and future extraction.",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--new-only", action="store_true", help="Only extract the two supplemental pet workbooks.")
    args = parser.parse_args()

    args.input_dir = args.input_dir.resolve()
    args.output_dir = args.output_dir.resolve()

    if args.new_only:
        mega = extract_mega_test_workbook(args.input_dir)
        value_calc = extract_pet_value_workbook(args.input_dir)
        raw_summaries = []
        for workbook_name in [MEGA_TEST_WORKBOOK, PET_VALUE_WORKBOOK]:
            workbook_path = args.input_dir / workbook_name
            if workbook_path.exists():
                raw_summaries.append(dump_workbook_raw(workbook_path, args.output_dir))
        write_json(args.output_dir / "pet_battle_mega_test.json", {
            "source_file": mega["source_file"],
            "generated_at": now_iso(),
            "zone_rankings": mega.get("zone_rankings", []),
            "result_sections": mega.get("result_sections", []),
            "calculator_drop_breakdowns": mega.get("calculator_drop_breakdowns", []),
            "loot_prices": mega.get("loot_prices", []),
            "food_prices": mega.get("food_prices", []),
            "test_sheets": mega.get("test_sheets", []),
        })
        write_json(args.output_dir / "pet_value_calculator.json", {
            "source_file": value_calc["source_file"],
            "generated_at": now_iso(),
            "pet_accounting": value_calc.get("pet_accounting", []),
            "egg_price_table": value_calc.get("egg_price_table", []),
            "level_100_bonus_table": value_calc.get("level_100_bonus_table", []),
            "leveling": value_calc.get("leveling", []),
            "valuation_formula": value_calc.get("valuation_formula", {}),
        })
        write_json(args.output_dir / "supplemental_workbook_manifest.json", {
            "generated_at": now_iso(),
            "files": {
                "pet_battle_mega_test.json": len(mega.get("calculator_drop_breakdowns", [])),
                "pet_value_calculator.json": len(value_calc.get("pet_accounting", [])),
            },
            "raw_workbook_dumps": raw_summaries,
        })
        print("Supplemental pet workbook extracts generated.")
        return

    chapter = extract_chapter_workbook(args.input_dir)
    master = extract_master_workbook(args.input_dir)
    mega = extract_mega_test_workbook(args.input_dir)
    value_calc = extract_pet_value_workbook(args.input_dir)

    raw_summaries = []
    for workbook_name in [MEGA_TEST_WORKBOOK, PET_VALUE_WORKBOOK]:
        workbook_path = args.input_dir / workbook_name
        if workbook_path.exists():
            raw_summaries.append(dump_workbook_raw(workbook_path, args.output_dir))

    write_json(args.output_dir / "pet_species_stats.json", {
        "source_file": chapter["source_file"],
        "generated_at": now_iso(),
        "species": chapter["species"],
        "known_formulas": chapter["known_formulas"],
    })
    write_json(args.output_dir / "pet_acquisition.json", {
        "source_file": chapter["source_file"],
        "generated_at": now_iso(),
        "acquisition": chapter["acquisition"],
    })
    write_json(args.output_dir / "pet_rarity.json", {
        "source_file": chapter["source_file"],
        "generated_at": now_iso(),
        "rarity": chapter["rarity"],
    })
    write_json(args.output_dir / "pet_mastery.json", {
        "source_file": chapter["source_file"],
        "generated_at": now_iso(),
        "pet_mastery": chapter["pet_mastery"],
        "concurrent_slots": chapter["concurrent_slots"],
        "loot_chance_breakpoints": chapter["loot_chance_breakpoints"],
        "stamina_drain_by_quality": chapter["stamina_drain_by_quality"],
    })
    write_json(args.output_dir / "pet_battle_zones.json", {
        "source_file": master["source_file"],
        "generated_at": now_iso(),
        "zones": master["zone_pet_battle_records"],
    })
    write_json(args.output_dir / "pet_battle_prices.json", {
        "source_file": master["source_file"],
        "generated_at": now_iso(),
        "prices": master["prices"],
        "food_prices": master["food_prices"],
    })
    write_json(args.output_dir / "pet_battle_expected_drops.json", {
        "source_file": master["source_file"],
        "generated_at": now_iso(),
        "expected_drops": master["expected_drops"],
    })
    write_json(args.output_dir / "pet_battle_mega_test.json", {
        "source_file": mega["source_file"],
        "generated_at": now_iso(),
        "zone_rankings": mega.get("zone_rankings", []),
        "result_sections": mega.get("result_sections", []),
        "calculator_drop_breakdowns": mega.get("calculator_drop_breakdowns", []),
        "loot_prices": mega.get("loot_prices", []),
        "food_prices": mega.get("food_prices", []),
        "test_sheets": mega.get("test_sheets", []),
    })
    write_json(args.output_dir / "pet_value_calculator.json", {
        "source_file": value_calc["source_file"],
        "generated_at": now_iso(),
        "pet_accounting": value_calc.get("pet_accounting", []),
        "egg_price_table": value_calc.get("egg_price_table", []),
        "level_100_bonus_table": value_calc.get("level_100_bonus_table", []),
        "leveling": value_calc.get("leveling", []),
        "valuation_formula": value_calc.get("valuation_formula", {}),
    })

    manifest = build_manifest(chapter, master, mega, value_calc, raw_summaries, args.output_dir)
    write_json(args.output_dir / "manifest.json", manifest)

    print("Pet workbook extracts generated.")
    print(json.dumps(manifest["files"], indent=2))


if __name__ == "__main__":
    main()
